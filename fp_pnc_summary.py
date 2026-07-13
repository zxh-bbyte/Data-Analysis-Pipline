from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.worksheet import Worksheet


# =============================================================================
# CONFIGURATION -- user-adjustable parameters
# =============================================================================
# Input workbook produced by particle_fp_pipeline.py, and the PNC output path.
# Default to the example analysis folder next to this script; change for real use.
_SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_PATH = _SCRIPT_DIR / "analysis" / "FPs_summary.xlsx"
OUTPUT_PATH = _SCRIPT_DIR / "analysis" / "FPs_PNC_summary.xlsx"

# Transport efficiency (TE) of the batch being processed. Process one TE batch
# per run and set this to that batch's measured transport efficiency.
TE = 0.4

# Note: the dilution factor is read per sample from the input workbook's
# Dilutionfactor column (parsed from each sample's "X<n>" tag upstream), not here.

# PNC (particles/mg) = counts * DilutionFactor * SECONDS_PER_MINUTE * CONSTANT_VOLUME_ML
#     / (TE * SAMPLE_FLOW_RATE_ML_MIN * ACQUISITION_TIME_SECONDS * SAMPLE_MASS_MG)
SECONDS_PER_MINUTE = 60
SAMPLE_FLOW_RATE_ML_MIN = 0.02       # nebulizer sample uptake rate (mL/min)
ACQUISITION_TIME_SECONDS = 150       # per-measurement acquisition time (s)
CONSTANT_VOLUME_ML = 50              # make-up (constant) volume (mL)
SAMPLE_MASS_MG = 20                  # weighed sample mass (mg)

PNC_LABEL = "PNCs(particles/mg)"
PREFERRED_SHEET_ORDER = ["smFPs", "mmFPs", "Total", "0.1FPs", "0.1-0.4FPs", "0.4-1FPs", "mainFPs"]
# =============================================================================
# End of configuration
# =============================================================================


def load_source_workbook():
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"Input workbook not found: {INPUT_PATH}")
    return load_workbook(INPUT_PATH, data_only=False)


def ordered_source_worksheets(source_workbook) -> list[Worksheet]:
    worksheet_map = {worksheet.title: worksheet for worksheet in source_workbook.worksheets}
    ordered_titles: list[str] = []

    for title in PREFERRED_SHEET_ORDER:
        if title in worksheet_map:
            ordered_titles.append(title)

    for worksheet in source_workbook.worksheets:
        if worksheet.title not in ordered_titles:
            ordered_titles.append(worksheet.title)

    return [worksheet_map[title] for title in ordered_titles]


def sheet_rows(worksheet: Worksheet) -> list[list[object]]:
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    while rows and all(cell is None for cell in rows[-1]):
        rows.pop()
    return rows


def is_row_layout(rows: list[list[object]]) -> bool:
    if not rows or not rows[0]:
        return False
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    return bool(headers) and headers[0] == "sampleID" and "Dilutionfactor" in headers


def is_column_layout(rows: list[list[object]]) -> bool:
    if not rows or not rows[0]:
        return False
    headers = rows[0]
    if str(headers[0]).strip() != "Element":
        return False

    sample_headers = [cell for cell in headers[1:] if cell is not None]
    return bool(sample_headers)


def ensure_output_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def coerce_numeric(value: object) -> object:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def build_row_layout_headers(source_headers: list[str]) -> list[str]:
    output_headers: list[str] = []
    for header in source_headers:
        output_headers.append(header)
        if header == "sampleID":
            output_headers.append("TE")
    return output_headers


def row_numeric_headers(headers: list[str], data_rows: list[list[object]]) -> set[str]:
    numeric_headers: set[str] = set()
    for column_index, header in enumerate(headers):
        if header in {"sampleID", "Dilutionfactor"}:
            continue

        values = [row[column_index] for row in data_rows if column_index < len(row) and row[column_index] is not None]
        if not values:
            continue
        if all(isinstance(value, (int, float)) for value in values):
            numeric_headers.add(header)
    return numeric_headers


def build_row_formula(
    counts_row: int,
    counts_col: int,
    pnc_row: int,
    te_col: int,
    dilution_col: int,
) -> str:
    counts_ref = f"{get_column_letter(counts_col)}{counts_row}"
    te_ref = f"${get_column_letter(te_col)}{pnc_row}"
    dilution_ref = f"${get_column_letter(dilution_col)}{pnc_row}"
    return (
        f"={counts_ref}*{dilution_ref}*{SECONDS_PER_MINUTE}*{CONSTANT_VOLUME_ML}"
        f"/({te_ref}*{SAMPLE_FLOW_RATE_ML_MIN}*{ACQUISITION_TIME_SECONDS}*{SAMPLE_MASS_MG})"
    )


def build_column_formula(
    counts_row: int,
    counts_col: int,
    te_row: int,
    dilution_factor: int | float,
) -> str:
    counts_ref = f"{get_column_letter(counts_col)}{counts_row}"
    te_ref = f"{get_column_letter(counts_col)}{te_row}"
    return (
        f"={counts_ref}*{dilution_factor}*{SECONDS_PER_MINUTE}*{CONSTANT_VOLUME_ML}"
        f"/({te_ref}*{SAMPLE_FLOW_RATE_ML_MIN}*{ACQUISITION_TIME_SECONDS}*{SAMPLE_MASS_MG})"
    )


def write_row(ws: Worksheet, row_index: int, values: Iterable[object]) -> None:
    for column_index, value in enumerate(values, start=1):
        ws.cell(row=row_index, column=column_index, value=value)


def build_dilution_map(source_workbook) -> dict[str, int | float]:
    dilution_map: dict[str, int | float] = {}

    for worksheet in source_workbook.worksheets:
        rows = sheet_rows(worksheet)
        if not is_row_layout(rows):
            continue

        headers = [str(cell).strip() for cell in rows[0]]
        sample_idx = headers.index("sampleID")
        dilution_idx = headers.index("Dilutionfactor")

        for row in rows[1:]:
            if not row or sample_idx >= len(row):
                continue
            sample_id = row[sample_idx]
            if sample_id is None:
                continue

            dilution_value = row[dilution_idx]
            if dilution_value is None:
                raise ValueError(
                    f"Missing Dilutionfactor for sample {sample_id!r} in sheet {worksheet.title!r}"
                )

            normalized_sample_id = str(sample_id).strip()
            dilution_value = coerce_numeric(dilution_value)

            existing_value = dilution_map.get(normalized_sample_id)
            if existing_value is not None and existing_value != dilution_value:
                raise ValueError(
                    f"Inconsistent Dilutionfactor for sample {normalized_sample_id!r}: "
                    f"{existing_value!r} vs {dilution_value!r}"
                )
            dilution_map[normalized_sample_id] = dilution_value

    if not dilution_map:
        raise ValueError("No row-layout sheets were found to build the dilution map.")

    return dilution_map


def write_row_layout_sheet(source_ws: Worksheet, target_ws: Worksheet) -> None:
    rows = sheet_rows(source_ws)
    headers = [str(cell).strip() for cell in rows[0]]
    data_rows = rows[1:]
    output_headers = build_row_layout_headers(headers)
    numeric_headers = row_numeric_headers(headers, data_rows)

    write_row(target_ws, 1, output_headers)

    counts_row_lookup: dict[str, int] = {}
    for source_offset, row in enumerate(data_rows, start=2):
        if not any(cell is not None for cell in row):
            continue

        sample_id = row[headers.index("sampleID")]
        te_value = TE

        output_values: list[object] = []
        for header, value in zip(headers, row):
            output_values.append(coerce_numeric(value))
            if header == "sampleID":
                output_values.append(te_value)

        write_row(target_ws, source_offset, output_values)
        counts_row_lookup[str(sample_id).strip()] = source_offset

    counts_last_row = 1 + len(counts_row_lookup)
    label_row = counts_last_row + 2
    pnc_header_row = counts_last_row + 4

    target_ws.cell(row=label_row, column=1, value=PNC_LABEL)
    write_row(target_ws, pnc_header_row, output_headers)

    header_to_output_col = {header: index + 1 for index, header in enumerate(output_headers)}
    te_col = header_to_output_col["TE"]
    dilution_col = header_to_output_col["Dilutionfactor"]

    source_data_rows = [row for row in data_rows if any(cell is not None for cell in row)]
    for output_offset, row in enumerate(source_data_rows, start=1):
        pnc_row = pnc_header_row + output_offset
        sample_id = str(row[headers.index("sampleID")]).strip()
        te_value = TE
        counts_row = counts_row_lookup[sample_id]

        output_values: list[object] = []
        for header, value in zip(headers, row):
            if header == "sampleID":
                output_values.append(sample_id)
                output_values.append(te_value)
            elif header == "Dilutionfactor":
                output_values.append(coerce_numeric(value))
            elif header in numeric_headers:
                counts_col = header_to_output_col[header]
                output_values.append(
                    build_row_formula(
                        counts_row=counts_row,
                        counts_col=counts_col,
                        pnc_row=pnc_row,
                        te_col=te_col,
                        dilution_col=dilution_col,
                    )
                )
            else:
                output_values.append(value)

        write_row(target_ws, pnc_row, output_values)


def write_column_layout_sheet(
    source_ws: Worksheet,
    target_ws: Worksheet,
    dilution_map: dict[str, int | float],
) -> None:
    rows = sheet_rows(source_ws)
    headers = [coerce_numeric(cell) for cell in rows[0]]
    data_rows = rows[1:]

    for row_index, row in enumerate(rows, start=1):
        write_row(target_ws, row_index, [coerce_numeric(value) for value in row])

    counts_last_row = len(rows)
    label_row = counts_last_row + 2
    pnc_header_row = counts_last_row + 4
    te_row = pnc_header_row + 1

    target_ws.cell(row=label_row, column=1, value=PNC_LABEL)
    write_row(target_ws, pnc_header_row, headers)

    target_ws.cell(row=te_row, column=1, value="TE")
    for column_index in range(2, len(headers) + 1):
        target_ws.cell(row=te_row, column=column_index, value=TE)

    for data_offset, row in enumerate(data_rows, start=1):
        pnc_row = te_row + data_offset
        counts_row = 1 + data_offset
        target_ws.cell(row=pnc_row, column=1, value=row[0])

        for column_index, header in enumerate(headers[1:], start=2):
            sample_id = str(header).strip()
            if sample_id not in dilution_map:
                raise ValueError(f"Missing Dilutionfactor mapping for sample {sample_id!r}")

            dilution_factor = dilution_map[sample_id]
            target_ws.cell(
                row=pnc_row,
                column=column_index,
                value=build_column_formula(
                    counts_row=counts_row,
                    counts_col=column_index,
                    te_row=te_row,
                    dilution_factor=dilution_factor,
                ),
            )


def write_sheet(
    source_ws: Worksheet,
    target_ws: Worksheet,
    dilution_map: dict[str, int | float],
) -> None:
    rows = sheet_rows(source_ws)
    if is_row_layout(rows):
        write_row_layout_sheet(source_ws, target_ws)
        return
    if is_column_layout(rows):
        write_column_layout_sheet(source_ws, target_ws, dilution_map=dilution_map)
        return

    headers = rows[0] if rows else []
    raise ValueError(
        f"Unrecognized sheet layout for {source_ws.title!r}. Headers: {headers!r}"
    )


def main() -> None:
    source_workbook = load_source_workbook()
    dilution_map = build_dilution_map(source_workbook)
    source_worksheets = ordered_source_worksheets(source_workbook)

    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)
    output_workbook.calculation = CalcProperties(
        calcMode="auto",
        fullCalcOnLoad=True,
        forceFullCalc=True,
    )

    for source_ws in source_worksheets:
        target_ws = output_workbook.create_sheet(title=source_ws.title)
        write_sheet(source_ws, target_ws, dilution_map=dilution_map)

    ensure_output_dir(OUTPUT_PATH)
    output_workbook.save(OUTPUT_PATH)
    print(f"PNC workbook written to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
