from __future__ import annotations

import math
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.worksheet import Worksheet


# =============================================================================
# CONFIGURATION -- user-adjustable parameters
# =============================================================================
# Folder of per-measurement particle CSVs (output of extract_particles.py).
_SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_DIR = _SCRIPT_DIR / "Particle"
OUTPUT_DIR = _SCRIPT_DIR / "analysis" / "Toxic"
SUMMARY_OUTPUT_PATH = OUTPUT_DIR / "ToxicFPs_summary.xlsx"
PNC_OUTPUT_PATH = OUTPUT_DIR / "ToxicFPs_PNC_summary.xlsx"

TOXIC_ELEMENTS = [
    "51V",
    "52Cr",
    "55Mn",
    "59Co",
    "60Ni",
    "63Cu",
    "66Zn",
    "75As",
    "111Cd",
    "120Sn",
    "121Sb",
    "138Ba",
    "205Tl",
    "208Pb",
]

RANGE_RULES: dict[str, dict[str, Any]] = {
    "smFPs": {"total_column": "TotalsmFPs", "exact": 1.0},
    "mmFPs": {
        "total_column": "TotalmmFPs",
        "min": 0.0,
        "max": 1.0,
        "include_min": False,
        "include_max": False,
    },
    "0.1FPs": {
        "total_column": "Total0.1FPs",
        "min": 0.0,
        "max": 0.1,
        "include_min": False,
        "include_max": True,
    },
    "0.1-0.4FPs": {
        "total_column": "Total0.1-0.4FPs",
        "min": 0.1,
        "max": 0.4,
        "include_min": False,
        "include_max": False,
    },
    "0.4-1FPs": {
        "total_column": "Total0.4-1FPs",
        "min": 0.4,
        "max": 1.0,
        "include_min": True,
        "include_max": False,
    },
}

TOTAL_TOTAL_COLUMN = "TotalFPs"
PNC_LABEL = "PNCs(particles/mg)"

# Transport efficiency (TE) of the batch being processed. Process one TE batch
# per run and set this to that batch's measured transport efficiency.
TE = 0.4

# PNC (particles/mg) = counts * DilutionFactor * SECONDS_PER_MINUTE * CONSTANT_VOLUME_ML
#     / (TE * SAMPLE_FLOW_RATE_ML_MIN * ACQUISITION_TIME_SECONDS * SAMPLE_MASS_MG)
# DilutionFactor is parsed per sample from the "X<n>" tag in the sample name
# (e.g. CFPP2X1000 -> 1000, EAF10X5000 -> 5000); it is not a single global value.
SECONDS_PER_MINUTE = 60
SAMPLE_FLOW_RATE_ML_MIN = 0.02       # nebulizer sample uptake rate (mL/min)
ACQUISITION_TIME_SECONDS = 150       # per-measurement acquisition time (s)
CONSTANT_VOLUME_ML = 50              # make-up (constant) volume (mL)
SAMPLE_MASS_MG = 20                  # weighed sample mass (mg)

# The optional trailing "-<n>x" is ignored; dilution comes from the sample name.
FILENAME_PATTERN = re.compile(
    r"^particle_(?P<sample_id>.+?)-(?P<measurement>\d+)(?:-\d+(?:\.\d+)?x)?\.csv$",
    re.IGNORECASE,
)

SUMMARY_SHEET_ORDER = [
    "smFPs_all",
    "mmFPs_all",
    "Total_all",
    "0.1FPs_all",
    "0.1-0.4FPs_all",
    "0.4-1FPs_all",
    "mainFPs_all",
    "smFPs_toxic",
    "mmFPs_toxic",
    "Total_toxic",
    "0.1FPs_toxic",
    "0.1-0.4FPs_toxic",
    "0.4-1FPs_toxic",
    "mainFPs_toxic",
    "mm_toxic_sum_measurement",
    "mm_toxic_sum_sample",
]

PNC_SHEET_ORDER = [
    "smFPs_all",
    "mmFPs_all",
    "Total_all",
    "0.1FPs_all",
    "0.1-0.4FPs_all",
    "0.4-1FPs_all",
    "mainFPs_all",
    "smFPs_toxic",
    "mmFPs_toxic",
    "Total_toxic",
    "0.1FPs_toxic",
    "0.1-0.4FPs_toxic",
    "0.4-1FPs_toxic",
    "mainFPs_toxic",
]
# =============================================================================
# End of configuration
# =============================================================================


@dataclass(frozen=True)
class ParticleFile:
    path: Path
    file_name: str
    sample_id: str
    measurement_index: int
    dilution_factor: int | float
    dilution_label: str
    sample_prefix: str
    sample_number: int

    @property
    def measurement_sample_id(self) -> str:
        return f"{self.sample_id}-{self.measurement_index}"

    @property
    def measurement_id(self) -> str:
        return f"{self.sample_id}-{self.measurement_index}-{self.dilution_label}"

    @property
    def group_key(self) -> tuple[str, int | float]:
        return self.sample_id, self.dilution_factor

    @property
    def sort_key(self) -> tuple[str, int, str, float, int]:
        return (
            self.sample_prefix.lower(),
            self.sample_number,
            self.sample_id.lower(),
            float(self.dilution_factor),
            self.measurement_index,
        )


@dataclass
class MeasurementData:
    meta: ParticleFile
    toxic_particles_df: pd.DataFrame
    cleaned_toxic_particles_df: pd.DataFrame
    mm_toxic_fraction_sums: pd.Series


def split_sample_for_sort(sample_id: str) -> tuple[str, int]:
    matched = re.match(r"^(?P<prefix>[A-Za-z]+)(?P<number>\d+)$", sample_id)
    if matched:
        return matched.group("prefix"), int(matched.group("number"))
    return sample_id, -1


def parse_dilution_from_sample(sample_id: str) -> int:
    """Dilution factor from the sample name's trailing "X<n>" tag (default 1)."""
    matched = re.search(r"[Xx](\d+)$", sample_id)
    return int(matched.group(1)) if matched else 1


def strip_dilution_tag(sample_id: str) -> str:
    """Drop the trailing "X<n>" dilution tag from a sample name (CFPP2X1000 -> CFPP2)."""
    return re.sub(r"[Xx]\d+$", "", sample_id)


def round_half_up(value: float) -> int:
    return int(math.floor(value + 0.5))


def ensure_output_dir(directory: Path) -> None:
    directory.mkdir(parents=True, exist_ok=True)


def discover_particle_files(input_dir: Path) -> list[ParticleFile]:
    files: list[ParticleFile] = []
    for file_path in input_dir.glob("particle_*.csv"):
        matched = FILENAME_PATTERN.match(file_path.name)
        if not matched:
            raise ValueError(f"Unable to parse file name: {file_path.name}")

        raw_sample_id = matched.group("sample_id")
        dilution_factor = parse_dilution_from_sample(raw_sample_id)
        sample_id = strip_dilution_tag(raw_sample_id)
        prefix, number = split_sample_for_sort(sample_id)
        files.append(
            ParticleFile(
                path=file_path,
                file_name=file_path.name,
                sample_id=sample_id,
                measurement_index=int(matched.group("measurement")),
                dilution_factor=dilution_factor,
                dilution_label=f"{dilution_factor}x",
                sample_prefix=prefix,
                sample_number=number,
            )
        )

    files.sort(key=lambda item: item.sort_key)
    if not files:
        raise FileNotFoundError(f"No particle CSV files found under: {input_dir}")
    return files


def read_particle_dataframe(file_path: Path, expected_columns: list[str] | None = None) -> pd.DataFrame:
    dataframe = pd.read_csv(file_path)
    if "embedding" in dataframe.columns:
        dataframe = dataframe.drop(columns=["embedding"])

    dataframe = dataframe.apply(pd.to_numeric, errors="coerce")

    if expected_columns is None:
        return dataframe

    missing_columns = [column for column in expected_columns if column not in dataframe.columns]
    extra_columns = [column for column in dataframe.columns if column not in expected_columns]
    if missing_columns or extra_columns:
        raise ValueError(
            f"Unexpected columns in {file_path.name}; "
            f"missing={missing_columns}, extra={extra_columns}"
        )
    return dataframe[expected_columns]


def normalize_particle_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Fill blanks in the raw particle table with 0 for whole-row conditions."""
    return dataframe.fillna(0)


def validate_toxic_columns(all_columns: list[str]) -> list[str]:
    present_columns = [column for column in TOXIC_ELEMENTS if column in all_columns]
    missing_columns = [column for column in TOXIC_ELEMENTS if column not in all_columns]
    if missing_columns:
        print(f"Warning: toxic elements not present in the data, skipping: {missing_columns}")
    if not present_columns:
        raise ValueError("None of the configured TOXIC_ELEMENTS are present in the particle files.")
    return present_columns


def build_range_mask(dataframe: pd.DataFrame, rule: dict[str, Any]) -> pd.DataFrame:
    if "exact" in rule:
        return dataframe.eq(rule["exact"])

    min_value = rule.get("min")
    max_value = rule.get("max")
    include_min = rule.get("include_min", False)
    include_max = rule.get("include_max", False)

    mask = pd.DataFrame(True, index=dataframe.index, columns=dataframe.columns)
    if min_value is not None:
        mask &= dataframe.ge(min_value) if include_min else dataframe.gt(min_value)
    if max_value is not None:
        mask &= dataframe.le(max_value) if include_max else dataframe.lt(max_value)
    return mask


def create_measurement_record(meta: ParticleFile, counts: pd.Series, total_column: str) -> dict[str, Any]:
    record: dict[str, Any] = {
        "_sample_id": meta.sample_id,
        "_sample_prefix": meta.sample_prefix.lower(),
        "_sample_number": meta.sample_number,
        "_dilution_sort": float(meta.dilution_factor),
        "_measurement_index": meta.measurement_index,
        "sampleID": meta.measurement_sample_id,
        "Dilutionfactor": meta.dilution_factor,
        total_column: int(counts.sum()),
    }
    record.update(counts.astype(int).to_dict())
    return record


def create_average_record(
    measurement_group: pd.DataFrame,
    sample_id: str,
    dilution_factor: int | float,
    total_column: str,
    element_columns: list[str],
) -> dict[str, Any]:
    average_counts = measurement_group[element_columns].mean(axis=0)
    rounded_counts = average_counts.apply(round_half_up).astype(int)

    record: dict[str, Any] = {
        "_sample_id": sample_id,
        "_sample_prefix": measurement_group.iloc[0]["_sample_prefix"],
        "_sample_number": int(measurement_group.iloc[0]["_sample_number"]),
        "_dilution_sort": float(dilution_factor),
        "_measurement_index": math.inf,
        "sampleID": sample_id,
        "Dilutionfactor": dilution_factor,
        total_column: int(rounded_counts.sum()),
    }
    record.update(rounded_counts.to_dict())
    return record


def finalize_average_table(
    records: list[dict[str, Any]],
    total_column: str,
    element_columns: list[str],
) -> pd.DataFrame:
    measurement_df = pd.DataFrame(records)
    measurement_df = measurement_df.sort_values(
        by=[
            "_sample_prefix",
            "_sample_number",
            "_sample_id",
            "_dilution_sort",
            "_measurement_index",
        ],
        kind="stable",
    ).reset_index(drop=True)

    average_records: list[dict[str, Any]] = []
    group_columns = [
        "_sample_prefix",
        "_sample_number",
        "_sample_id",
        "_dilution_sort",
        "Dilutionfactor",
    ]
    for _, measurement_group in measurement_df.groupby(group_columns, sort=False):
        sample_id = str(measurement_group.iloc[0]["_sample_id"])
        dilution_factor = measurement_group.iloc[0]["Dilutionfactor"]
        average_records.append(
            create_average_record(
                measurement_group=measurement_group,
                sample_id=sample_id,
                dilution_factor=dilution_factor,
                total_column=total_column,
                element_columns=element_columns,
            )
        )

    output_columns = ["sampleID", "Dilutionfactor", total_column] + element_columns
    average_df = pd.DataFrame(average_records)
    average_df = average_df.sort_values(
        by=["_sample_prefix", "_sample_number", "_sample_id", "_dilution_sort"],
        kind="stable",
    ).reset_index(drop=True)
    return average_df[output_columns]


def build_total_summary_table(
    sm_table: pd.DataFrame,
    mm_table: pd.DataFrame,
    sm_total_column: str,
    mm_total_column: str,
) -> pd.DataFrame:
    key_columns = ["sampleID", "Dilutionfactor"]
    sm_element_columns = [column for column in sm_table.columns if column not in key_columns + [sm_total_column]]
    mm_element_columns = [column for column in mm_table.columns if column not in key_columns + [mm_total_column]]
    if sm_element_columns != mm_element_columns:
        raise ValueError("smFPs and mmFPs element columns do not match for Total table.")

    merged = sm_table.merge(
        mm_table,
        on=key_columns,
        how="inner",
        suffixes=("_sm", "_mm"),
        validate="one_to_one",
    )
    if len(merged) != len(sm_table) or len(merged) != len(mm_table):
        raise ValueError("smFPs and mmFPs sample sets do not match for Total table.")

    total_table = merged[key_columns].copy()
    total_table[TOTAL_TOTAL_COLUMN] = merged[sm_total_column] + merged[mm_total_column]
    for element in sm_element_columns:
        total_table[element] = merged[f"{element}_sm"] + merged[f"{element}_mm"]
    return total_table


def clean_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    return dataframe.loc[~dataframe.eq(1).any(axis=1)].copy()


def compute_dominant_element_counts(dataframe: pd.DataFrame, element_columns: list[str]) -> pd.Series:
    if dataframe.empty:
        return pd.Series(0, index=element_columns, dtype=int)
    dominant_elements = dataframe[element_columns].fillna(0).idxmax(axis=1)
    return dominant_elements.value_counts().reindex(element_columns, fill_value=0).astype(int)


def build_average_main_table(measurement_data: list[MeasurementData], element_columns: list[str]) -> pd.DataFrame:
    grouped_counts: dict[tuple[str, int | float], list[pd.Series]] = defaultdict(list)
    grouped_metadata: dict[tuple[str, int | float], list[ParticleFile]] = defaultdict(list)

    for item in measurement_data:
        counts = compute_dominant_element_counts(item.cleaned_toxic_particles_df, element_columns)
        grouped_counts[item.meta.group_key].append(counts)
        grouped_metadata[item.meta.group_key].append(item.meta)

    average_only = pd.DataFrame(index=element_columns)
    average_only.index.name = "Element"
    ordered_group_keys = sorted(grouped_metadata, key=lambda key: grouped_metadata[key][0].sort_key)
    for group_key in ordered_group_keys:
        average_series = pd.concat(grouped_counts[group_key], axis=1).mean(axis=1)
        average_series = average_series.apply(round_half_up).astype(int)
        average_only[group_key[0]] = average_series.reindex(element_columns, fill_value=0).astype(int)

    return average_only.reset_index()


def prepare_measurement_data(
    particle_files: list[ParticleFile],
) -> tuple[list[MeasurementData], list[str], list[str]]:
    measurement_data: list[MeasurementData] = []
    all_columns: list[str] | None = None
    toxic_columns: list[str] | None = None

    for particle_file in particle_files:
        dataframe = read_particle_dataframe(particle_file.path, expected_columns=all_columns)
        if all_columns is None:
            all_columns = dataframe.columns.tolist()
            toxic_columns = validate_toxic_columns(all_columns)

        filled_dataframe = normalize_particle_dataframe(dataframe)
        toxic_fraction_sum = filled_dataframe[toxic_columns].sum(axis=1)
        toxic_particles_df = filled_dataframe.loc[toxic_fraction_sum.gt(0), all_columns].copy()
        cleaned_toxic_particles_df = clean_dataframe(toxic_particles_df)

        mm_particle_mask = toxic_particles_df.lt(1).all(axis=1) & toxic_particles_df.gt(0).any(axis=1)
        mm_toxic_fraction_sums = toxic_particles_df.loc[mm_particle_mask, toxic_columns].sum(axis=1)

        measurement_data.append(
            MeasurementData(
                meta=particle_file,
                toxic_particles_df=toxic_particles_df,
                cleaned_toxic_particles_df=cleaned_toxic_particles_df,
                mm_toxic_fraction_sums=mm_toxic_fraction_sums.astype(float),
            )
        )

    if all_columns is None or toxic_columns is None:
        raise ValueError("No particle data were loaded.")
    return measurement_data, all_columns, toxic_columns


def compute_scope_average_tables(
    measurement_data: list[MeasurementData],
    element_columns: list[str],
) -> dict[str, pd.DataFrame]:
    tables: dict[str, pd.DataFrame] = {}
    for rule_name, rule in RANGE_RULES.items():
        records: list[dict[str, Any]] = []
        for item in measurement_data:
            mask = build_range_mask(item.toxic_particles_df[element_columns], rule)
            counts = mask.sum(axis=0).reindex(element_columns, fill_value=0).astype(int)
            records.append(
                create_measurement_record(
                    meta=item.meta,
                    counts=counts,
                    total_column=rule["total_column"],
                )
            )
        tables[rule_name] = finalize_average_table(
            records=records,
            total_column=rule["total_column"],
            element_columns=element_columns,
        )
    return tables


def coerce_int_like(dataframe: pd.DataFrame) -> pd.DataFrame:
    coerced = dataframe.copy()
    for column in coerced.columns:
        if pd.api.types.is_float_dtype(coerced[column]):
            if coerced[column].dropna().map(float.is_integer).all():
                coerced[column] = coerced[column].astype("Int64").astype(object)
    return coerced


def build_summary_tables(
    measurement_data: list[MeasurementData],
    all_columns: list[str],
    toxic_columns: list[str],
) -> dict[str, pd.DataFrame]:
    all_scope_tables = compute_scope_average_tables(measurement_data, all_columns)
    toxic_scope_tables = compute_scope_average_tables(measurement_data, toxic_columns)

    summary_tables: dict[str, pd.DataFrame] = {
        "smFPs_all": all_scope_tables["smFPs"],
        "mmFPs_all": all_scope_tables["mmFPs"],
        "Total_all": build_total_summary_table(
            all_scope_tables["smFPs"],
            all_scope_tables["mmFPs"],
            RANGE_RULES["smFPs"]["total_column"],
            RANGE_RULES["mmFPs"]["total_column"],
        ),
        "0.1FPs_all": all_scope_tables["0.1FPs"],
        "0.1-0.4FPs_all": all_scope_tables["0.1-0.4FPs"],
        "0.4-1FPs_all": all_scope_tables["0.4-1FPs"],
        "mainFPs_all": build_average_main_table(measurement_data, all_columns),
        "smFPs_toxic": toxic_scope_tables["smFPs"],
        "mmFPs_toxic": toxic_scope_tables["mmFPs"],
        "Total_toxic": build_total_summary_table(
            toxic_scope_tables["smFPs"],
            toxic_scope_tables["mmFPs"],
            RANGE_RULES["smFPs"]["total_column"],
            RANGE_RULES["mmFPs"]["total_column"],
        ),
        "0.1FPs_toxic": toxic_scope_tables["0.1FPs"],
        "0.1-0.4FPs_toxic": toxic_scope_tables["0.1-0.4FPs"],
        "0.4-1FPs_toxic": toxic_scope_tables["0.4-1FPs"],
        "mainFPs_toxic": build_average_main_table(measurement_data, toxic_columns),
    }
    return {sheet_name: coerce_int_like(dataframe) for sheet_name, dataframe in summary_tables.items()}


def rows_sort_key(group_key: tuple[str, int | float]) -> tuple[str, int, str, float]:
    sample_id = group_key[0]
    prefix, number = split_sample_for_sort(sample_id)
    return prefix.lower(), number, sample_id.lower(), float(group_key[1])


def build_mm_toxic_sum_tables(measurement_data: list[MeasurementData]) -> tuple[pd.DataFrame, pd.DataFrame]:
    measurement_rows: list[dict[str, Any]] = []
    sample_groups: dict[tuple[str, int | float], list[dict[str, Any]]] = defaultdict(list)
    pooled_sums_by_group: dict[tuple[str, int | float], list[pd.Series]] = defaultdict(list)

    for item in measurement_data:
        particle_count = int(item.mm_toxic_fraction_sums.shape[0])
        measurement_row = {
            "_sample_prefix": item.meta.sample_prefix.lower(),
            "_sample_number": item.meta.sample_number,
            "_sample_id": item.meta.sample_id,
            "_dilution_sort": float(item.meta.dilution_factor),
            "_measurement_index": item.meta.measurement_index,
            "measurementID": item.meta.measurement_id,
            "sampleID": item.meta.sample_id,
            "Dilutionfactor": item.meta.dilution_factor,
            "particleCount": particle_count,
            "meanToxicFractionSum": float(item.mm_toxic_fraction_sums.mean()) if particle_count else 0.0,
        }
        measurement_rows.append(measurement_row)
        sample_groups[item.meta.group_key].append(measurement_row)
        pooled_sums_by_group[item.meta.group_key].append(item.mm_toxic_fraction_sums)

    measurement_df = pd.DataFrame(measurement_rows).sort_values(
        by=[
            "_sample_prefix",
            "_sample_number",
            "_sample_id",
            "_dilution_sort",
            "_measurement_index",
        ],
        kind="stable",
    ).reset_index(drop=True)
    measurement_output = measurement_df[
        ["measurementID", "sampleID", "Dilutionfactor", "particleCount", "meanToxicFractionSum"]
    ].copy()

    sample_rows: list[dict[str, Any]] = []
    for group_key in sorted(sample_groups, key=rows_sort_key):
        rows = sample_groups[group_key]
        pooled_parts = [series for series in pooled_sums_by_group[group_key] if not series.empty]
        pooled_mean = float(pd.concat(pooled_parts, axis=0).mean()) if pooled_parts else 0.0

        sample_rows.append(
            {
                "sampleID": group_key[0],
                "Dilutionfactor": group_key[1],
                "measurementCount": len(rows),
                "particleCount": int(sum(int(row["particleCount"]) for row in rows)),
                "meanFromMeasurements": float(sum(float(row["meanToxicFractionSum"]) for row in rows) / len(rows)),
                "meanFromPooledParticles": pooled_mean,
            }
        )

    sample_output = pd.DataFrame(sample_rows)
    return measurement_output, sample_output


def write_summary_workbook(
    summary_tables: dict[str, pd.DataFrame],
    mm_measurement_df: pd.DataFrame,
    mm_sample_df: pd.DataFrame,
) -> None:
    ensure_output_dir(OUTPUT_DIR)
    with pd.ExcelWriter(SUMMARY_OUTPUT_PATH, engine="openpyxl") as writer:
        for sheet_name in SUMMARY_SHEET_ORDER:
            if sheet_name == "mm_toxic_sum_measurement":
                mm_measurement_df.to_excel(writer, sheet_name=sheet_name, index=False)
            elif sheet_name == "mm_toxic_sum_sample":
                mm_sample_df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                summary_tables[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)


def load_source_workbook(path: Path):
    if not path.exists():
        raise FileNotFoundError(f"Input workbook not found: {path}")
    return load_workbook(path, data_only=False)


def ordered_source_worksheets(source_workbook) -> list[Worksheet]:
    worksheet_map = {worksheet.title: worksheet for worksheet in source_workbook.worksheets}
    missing_sheets = [sheet for sheet in PNC_SHEET_ORDER if sheet not in worksheet_map]
    if missing_sheets:
        raise ValueError(f"Missing expected sheets for PNC output: {missing_sheets}")
    return [worksheet_map[title] for title in PNC_SHEET_ORDER]


def sheet_rows(worksheet: Worksheet) -> list[list[object]]:
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    while rows and all(cell is None for cell in rows[-1]):
        rows.pop()
    return rows


def is_row_layout(rows: list[list[object]]) -> bool:
    if not rows or not rows[0]:
        return False
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    return bool(headers) and headers[0] == "sampleID" and "Dilutionfactor" in headers


def is_column_layout(rows: list[list[object]]) -> bool:
    if not rows or not rows[0]:
        return False
    headers = rows[0]
    if str(headers[0]).strip() != "Element":
        return False

    sample_headers = [cell for cell in headers[1:] if cell is not None]
    return bool(sample_headers)


def coerce_numeric(value: object) -> object:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def build_row_layout_headers(source_headers: list[str]) -> list[str]:
    output_headers: list[str] = []
    for header in source_headers:
        output_headers.append(header)
        if header == "sampleID":
            output_headers.append("TE")
    return output_headers


def row_numeric_headers(headers: list[str], data_rows: list[list[object]]) -> set[str]:
    numeric_headers: set[str] = set()
    for column_index, header in enumerate(headers):
        if header in {"sampleID", "Dilutionfactor"}:
            continue
        values = [
            row[column_index]
            for row in data_rows
            if column_index < len(row) and row[column_index] is not None
        ]
        if values and all(isinstance(value, (int, float)) for value in values):
            numeric_headers.add(header)
    return numeric_headers


def build_row_formula(
    counts_row: int,
    counts_col: int,
    pnc_row: int,
    te_col: int,
    dilution_col: int,
) -> str:
    counts_ref = f"{get_column_letter(counts_col)}{counts_row}"
    te_ref = f"${get_column_letter(te_col)}{pnc_row}"
    dilution_ref = f"${get_column_letter(dilution_col)}{pnc_row}"
    return (
        f"={counts_ref}*{dilution_ref}*{SECONDS_PER_MINUTE}*{CONSTANT_VOLUME_ML}"
        f"/({te_ref}*{SAMPLE_FLOW_RATE_ML_MIN}*{ACQUISITION_TIME_SECONDS}*{SAMPLE_MASS_MG})"
    )


def build_column_formula(
    counts_row: int,
    counts_col: int,
    te_row: int,
    dilution_factor: int | float,
) -> str:
    counts_ref = f"{get_column_letter(counts_col)}{counts_row}"
    te_ref = f"{get_column_letter(counts_col)}{te_row}"
    return (
        f"={counts_ref}*{dilution_factor}*{SECONDS_PER_MINUTE}*{CONSTANT_VOLUME_ML}"
        f"/({te_ref}*{SAMPLE_FLOW_RATE_ML_MIN}*{ACQUISITION_TIME_SECONDS}*{SAMPLE_MASS_MG})"
    )


def write_row(ws: Worksheet, row_index: int, values: Iterable[object]) -> None:
    for column_index, value in enumerate(values, start=1):
        ws.cell(row=row_index, column=column_index, value=value)


def build_dilution_map(source_workbook) -> dict[str, int | float]:
    dilution_map: dict[str, int | float] = {}
    for worksheet in ordered_source_worksheets(source_workbook):
        rows = sheet_rows(worksheet)
        if not is_row_layout(rows):
            continue

        headers = [str(cell).strip() for cell in rows[0]]
        sample_idx = headers.index("sampleID")
        dilution_idx = headers.index("Dilutionfactor")
        for row in rows[1:]:
            if not row or sample_idx >= len(row):
                continue
            sample_id = row[sample_idx]
            if sample_id is None:
                continue

            dilution_value = row[dilution_idx]
            if dilution_value is None:
                raise ValueError(
                    f"Missing Dilutionfactor for sample {sample_id!r} in sheet {worksheet.title!r}"
                )

            normalized_sample_id = str(sample_id).strip()
            dilution_value = coerce_numeric(dilution_value)
            existing_value = dilution_map.get(normalized_sample_id)
            if existing_value is not None and existing_value != dilution_value:
                raise ValueError(
                    f"Inconsistent Dilutionfactor for sample {normalized_sample_id!r}: "
                    f"{existing_value!r} vs {dilution_value!r}"
                )
            dilution_map[normalized_sample_id] = dilution_value

    if not dilution_map:
        raise ValueError("No row-layout sheets were found to build the dilution map.")
    return dilution_map


def write_row_layout_sheet(source_ws: Worksheet, target_ws: Worksheet) -> None:
    rows = sheet_rows(source_ws)
    headers = [str(cell).strip() for cell in rows[0]]
    data_rows = rows[1:]
    output_headers = build_row_layout_headers(headers)
    numeric_headers = row_numeric_headers(headers, data_rows)

    write_row(target_ws, 1, output_headers)

    counts_row_lookup: dict[str, int] = {}
    for source_offset, row in enumerate(data_rows, start=2):
        if not any(cell is not None for cell in row):
            continue

        sample_id = row[headers.index("sampleID")]
        te_value = TE

        output_values: list[object] = []
        for header, value in zip(headers, row):
            output_values.append(coerce_numeric(value))
            if header == "sampleID":
                output_values.append(te_value)

        write_row(target_ws, source_offset, output_values)
        counts_row_lookup[str(sample_id).strip()] = source_offset

    counts_last_row = 1 + len(counts_row_lookup)
    label_row = counts_last_row + 2
    pnc_header_row = counts_last_row + 4

    target_ws.cell(row=label_row, column=1, value=PNC_LABEL)
    write_row(target_ws, pnc_header_row, output_headers)

    header_to_output_col = {header: index + 1 for index, header in enumerate(output_headers)}
    te_col = header_to_output_col["TE"]
    dilution_col = header_to_output_col["Dilutionfactor"]

    source_data_rows = [row for row in data_rows if any(cell is not None for cell in row)]
    for output_offset, row in enumerate(source_data_rows, start=1):
        pnc_row = pnc_header_row + output_offset
        sample_id = str(row[headers.index("sampleID")]).strip()
        te_value = TE
        counts_row = counts_row_lookup[sample_id]

        output_values: list[object] = []
        for header, value in zip(headers, row):
            if header == "sampleID":
                output_values.append(sample_id)
                output_values.append(te_value)
            elif header == "Dilutionfactor":
                output_values.append(coerce_numeric(value))
            elif header in numeric_headers:
                counts_col = header_to_output_col[header]
                output_values.append(
                    build_row_formula(
                        counts_row=counts_row,
                        counts_col=counts_col,
                        pnc_row=pnc_row,
                        te_col=te_col,
                        dilution_col=dilution_col,
                    )
                )
            else:
                output_values.append(value)

        write_row(target_ws, pnc_row, output_values)


def write_column_layout_sheet(
    source_ws: Worksheet,
    target_ws: Worksheet,
    dilution_map: dict[str, int | float],
) -> None:
    rows = sheet_rows(source_ws)
    headers = [coerce_numeric(cell) for cell in rows[0]]
    data_rows = rows[1:]

    for row_index, row in enumerate(rows, start=1):
        write_row(target_ws, row_index, [coerce_numeric(value) for value in row])

    counts_last_row = len(rows)
    label_row = counts_last_row + 2
    pnc_header_row = counts_last_row + 4
    te_row = pnc_header_row + 1

    target_ws.cell(row=label_row, column=1, value=PNC_LABEL)
    write_row(target_ws, pnc_header_row, headers)

    target_ws.cell(row=te_row, column=1, value="TE")
    for column_index in range(2, len(headers) + 1):
        target_ws.cell(row=te_row, column=column_index, value=TE)

    for data_offset, row in enumerate(data_rows, start=1):
        pnc_row = te_row + data_offset
        counts_row = 1 + data_offset
        target_ws.cell(row=pnc_row, column=1, value=row[0])

        for column_index, header in enumerate(headers[1:], start=2):
            sample_id = str(header).strip()
            if sample_id not in dilution_map:
                raise ValueError(f"Missing Dilutionfactor mapping for sample {sample_id!r}")

            dilution_factor = dilution_map[sample_id]
            target_ws.cell(
                row=pnc_row,
                column=column_index,
                value=build_column_formula(
                    counts_row=counts_row,
                    counts_col=column_index,
                    te_row=te_row,
                    dilution_factor=dilution_factor,
                ),
            )


def write_sheet(
    source_ws: Worksheet,
    target_ws: Worksheet,
    dilution_map: dict[str, int | float],
) -> None:
    rows = sheet_rows(source_ws)
    if is_row_layout(rows):
        write_row_layout_sheet(source_ws, target_ws)
        return
    if is_column_layout(rows):
        write_column_layout_sheet(source_ws, target_ws, dilution_map=dilution_map)
        return
    headers = rows[0] if rows else []
    raise ValueError(f"Unrecognized sheet layout for {source_ws.title!r}. Headers: {headers!r}")


def write_pnc_workbook() -> None:
    source_workbook = load_source_workbook(SUMMARY_OUTPUT_PATH)
    dilution_map = build_dilution_map(source_workbook)
    source_worksheets = ordered_source_worksheets(source_workbook)

    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)
    output_workbook.calculation = CalcProperties(
        calcMode="auto",
        fullCalcOnLoad=True,
        forceFullCalc=True,
    )

    for source_ws in source_worksheets:
        target_ws = output_workbook.create_sheet(title=source_ws.title)
        write_sheet(source_ws, target_ws, dilution_map=dilution_map)

    ensure_output_dir(OUTPUT_DIR)
    output_workbook.save(PNC_OUTPUT_PATH)


def run_pipeline() -> None:
    particle_files = discover_particle_files(INPUT_DIR)
    measurement_data, all_columns, toxic_columns = prepare_measurement_data(particle_files)
    summary_tables = build_summary_tables(measurement_data, all_columns, toxic_columns)
    mm_measurement_df, mm_sample_df = build_mm_toxic_sum_tables(measurement_data)

    write_summary_workbook(summary_tables, mm_measurement_df, mm_sample_df)
    write_pnc_workbook()

    print(f"Toxic summary workbook written to: {SUMMARY_OUTPUT_PATH}")
    print(f"Toxic PNC workbook written to: {PNC_OUTPUT_PATH}")


if __name__ == "__main__":
    run_pipeline()
