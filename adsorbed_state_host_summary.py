"""Adsorbed-state host matrices, per sample (counts and PNC).

For adsorbed-state fingerprints (isotope fraction <= 0.1), this reports, per
sample and per adsorbed element, the particle counts split by host element (the
per-particle argmax element) and the matching particle number concentration
(PNC, particles/mg). Output is two per-sample host matrices, `host_counts` and
`host_pnc` (first column = sampleID). Works with any sample naming; it does not
depend on sample numbering or depth grouping.
"""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


# =============================================================================
# CONFIGURATION -- user-adjustable parameters
# =============================================================================
_SCRIPT_DIR = Path(__file__).resolve().parent
ANALYSIS_DIR = _SCRIPT_DIR / "analysis"
# Input workbook produced by fp_pnc_summary.py.
PNS_WORKBOOK_PATH = ANALYSIS_DIR / "FPs_PNC_summary.xlsx"
# Folder of per-measurement particle CSVs (output of extract_particles.py).
PARTICLE_DIR = _SCRIPT_DIR / "Particle"
OUTPUT_PATH = ANALYSIS_DIR / "0.1FPs_adsorbed_host_summary.xlsx"

# Worksheet in the PNC workbook holding the adsorbed-state (<=0.1) counts.
ADSORBED_SHEET_NAME = "0.1FPs"

# PNC scaling = DilutionFactor * SECONDS_PER_MINUTE * CONSTANT_VOLUME_ML
#     / (TE * SAMPLE_FLOW_RATE_ML_MIN * ACQUISITION_TIME_SECONDS * SAMPLE_MASS_MG)
SECONDS_PER_MINUTE = 60
SAMPLE_FLOW_RATE_ML_MIN = 0.02       # nebulizer sample uptake rate (mL/min)
ACQUISITION_TIME_SECONDS = 150       # per-measurement acquisition time (s)
CONSTANT_VOLUME_ML = 50              # make-up (constant) volume (mL)
SAMPLE_MASS_MG = 20                  # weighed sample mass (mg)
# =============================================================================
# End of configuration
# =============================================================================

PARTICLE_FILE_PATTERN = re.compile(
    r"^particle_(?P<sample_id>.+?)-(?P<measurement>\d+)(?:-\d+(?:\.\d+)?x)?\.csv$",
    re.IGNORECASE,
)


def strip_dilution_tag(sample_id: str) -> str:
    """Drop the trailing "X<n>" dilution tag from a sample name (CFPP2X1000 -> CFPP2)."""
    return re.sub(r"[Xx]\d+$", "", sample_id)


def load_top_counts_block(path: Path, sheet_name: str) -> pd.DataFrame:
    """Read the top counts block (sampleID / TE / Dilutionfactor / ...) of a sheet."""
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    workbook: Workbook = load_workbook(path, data_only=False, read_only=True)
    worksheet = workbook[sheet_name]

    rows: list[list[object]] = []
    data_started = False
    for row in worksheet.iter_rows(values_only=True):
        values = list(row)
        if not any(value is not None for value in values):
            if data_started:
                break
            continue
        data_started = True
        rows.append(values)

    if not rows:
        raise ValueError(f"No rows found in sheet {sheet_name!r}")

    header = rows[0]
    dataframe = pd.DataFrame(rows[1:], columns=header)
    dataframe = dataframe.loc[:, ~pd.isna(dataframe.columns)].copy()
    for column in ("TE", "Dilutionfactor"):
        if column not in dataframe.columns:
            raise ValueError(f"Sheet {sheet_name!r} has no {column!r} column.")
        dataframe[column] = pd.to_numeric(dataframe[column], errors="raise")
    dataframe["sampleID"] = dataframe["sampleID"].astype(str)
    return dataframe


def compute_scaling_by_sample(counts_df: pd.DataFrame) -> tuple[dict[str, float], list[str]]:
    """Per-sample PNC scaling factor, plus the sample order from the workbook."""
    scaling = (
        counts_df["Dilutionfactor"] * SECONDS_PER_MINUTE * CONSTANT_VOLUME_ML
        / (counts_df["TE"] * SAMPLE_FLOW_RATE_ML_MIN * ACQUISITION_TIME_SECONDS * SAMPLE_MASS_MG)
    )
    scaling_by_sample = dict(zip(counts_df["sampleID"], scaling.astype(float)))
    return scaling_by_sample, counts_df["sampleID"].tolist()


def collect_host_counts_by_sample(
    particle_dir: Path,
) -> tuple[list[str], dict[str, np.ndarray], dict[str, int]]:
    """Sum, per sample, an [adsorbed_element x host_element] particle-count matrix.

    For every particle where an element is in the adsorbed state (0 < v <= 0.1),
    the host is that particle's argmax element. Counts are summed over the
    sample's measurements; the measurement count is returned for averaging.
    """
    element_columns: list[str] | None = None
    host_sums: dict[str, np.ndarray] = {}
    measurement_counts: dict[str, int] = defaultdict(int)

    for file_path in sorted(particle_dir.glob("particle_*.csv")):
        matched = PARTICLE_FILE_PATTERN.match(file_path.name)
        if not matched:
            continue
        sample_id = strip_dilution_tag(matched.group("sample_id"))

        dataframe = pd.read_csv(file_path)
        if "embedding" in dataframe.columns:
            dataframe = dataframe.drop(columns=["embedding"])
        dataframe = dataframe.apply(pd.to_numeric, errors="coerce").fillna(0.0)
        if element_columns is None:
            element_columns = dataframe.columns.tolist()
        elif dataframe.columns.tolist() != element_columns:
            raise ValueError(f"Column mismatch in particle file: {file_path.name}")

        element_count = len(element_columns)
        array = dataframe.to_numpy(dtype=float)
        adsorbed_mask = (array > 0) & (array <= 0.1)
        host_indices = array.argmax(axis=1)

        matrix = np.zeros((element_count, element_count), dtype=float)
        for adsorbed_index in range(element_count):
            rows_mask = adsorbed_mask[:, adsorbed_index]
            if not rows_mask.any():
                continue
            matrix[adsorbed_index, :] += np.bincount(
                host_indices[rows_mask], minlength=element_count
            )

        host_sums.setdefault(sample_id, np.zeros((element_count, element_count), dtype=float))
        host_sums[sample_id] += matrix
        measurement_counts[sample_id] += 1

    if element_columns is None:
        raise ValueError(f"No particle CSV files found in: {particle_dir}")
    return element_columns, host_sums, measurement_counts


def build_host_tables(
    element_columns: list[str],
    host_sums: dict[str, np.ndarray],
    measurement_counts: dict[str, int],
    scaling_by_sample: dict[str, float],
    sample_order: list[str],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Two per-sample host matrices: particle counts, and the matching PNC.

    Counts are the per-sample mean over measurements (rounded half-up, as in the
    FP pipeline); PNC = count * per-sample scaling. One row per
    (sample, adsorbed element); host-element columns hold the value.
    """
    count_rows: list[dict[str, object]] = []
    pnc_rows: list[dict[str, object]] = []
    for sample_id in sample_order:
        if sample_id not in host_sums:
            continue
        scaling = scaling_by_sample.get(sample_id)
        if scaling is None:
            print(f"Warning: no TE/Dilutionfactor for sample {sample_id!r}, skipping.")
            continue

        counts = np.floor(host_sums[sample_id] / measurement_counts[sample_id] + 0.5)
        for adsorbed_index, adsorbed_element in enumerate(element_columns):
            total_count = int(counts[adsorbed_index, :].sum())
            if total_count <= 0:
                continue

            count_row: dict[str, object] = {
                "sampleID": sample_id,
                "adsorbed_element": adsorbed_element,
                "total_host_count": total_count,
            }
            pnc_row: dict[str, object] = {
                "sampleID": sample_id,
                "adsorbed_element": adsorbed_element,
                "total_host_pnc": float(total_count * scaling),
            }
            for host_index, host_element in enumerate(element_columns):
                cell_count = int(counts[adsorbed_index, host_index])
                count_row[host_element] = cell_count
                pnc_row[host_element] = float(cell_count * scaling)
            count_rows.append(count_row)
            pnc_rows.append(pnc_row)

    count_columns = ["sampleID", "adsorbed_element", "total_host_count"] + element_columns
    pnc_columns = ["sampleID", "adsorbed_element", "total_host_pnc"] + element_columns
    counts_df = pd.DataFrame(count_rows, columns=count_columns)
    pnc_df = pd.DataFrame(pnc_rows, columns=pnc_columns)

    sample_rank = {sample_id: index for index, sample_id in enumerate(sample_order)}
    for dataframe, total_column in ((counts_df, "total_host_count"), (pnc_df, "total_host_pnc")):
        if dataframe.empty:
            continue
        dataframe["_sample_rank"] = dataframe["sampleID"].map(sample_rank)
        dataframe.sort_values(
            by=["_sample_rank", total_column],
            ascending=[True, False],
            kind="stable",
            inplace=True,
        )
        dataframe.drop(columns=["_sample_rank"], inplace=True)
        dataframe.reset_index(drop=True, inplace=True)
    return counts_df, pnc_df


def main() -> None:
    counts_df = load_top_counts_block(PNS_WORKBOOK_PATH, ADSORBED_SHEET_NAME)
    scaling_by_sample, sample_order = compute_scaling_by_sample(counts_df)

    element_columns, host_sums, measurement_counts = collect_host_counts_by_sample(PARTICLE_DIR)
    host_counts_df, host_pnc_df = build_host_tables(
        element_columns=element_columns,
        host_sums=host_sums,
        measurement_counts=measurement_counts,
        scaling_by_sample=scaling_by_sample,
        sample_order=sample_order,
    )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        host_counts_df.to_excel(writer, sheet_name="host_counts", index=False)
        host_pnc_df.to_excel(writer, sheet_name="host_pnc", index=False)

    print(f"Adsorbed-state host matrices written to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
